import openpyxl

OLD_FILE = "/Users/mac/new/Diyxx/台式机核价专用免费版4.21xlsm.xlsm"
NEW_FILE = "/Users/mac/new/Diyxx/台式机核价专用免费版5.07.xlsm"

SHEETS_TO_CHECK = ["处理器", "主板", "内存", "硬盘", "显卡", "电源", "机箱", "显示器", "散热", "外设"]

def extract_prices(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    data = {}
    for sheet_name in SHEETS_TO_CHECK:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_data = {}
        for row in ws.iter_rows(min_row=2, max_row=10000):
            cells = {c.column: c.value for c in row if c.value is not None}
            model = cells.get(1)
            if not model or not str(model).strip():
                continue
            model = str(model).strip()
            recycle = float(cells.get(2, 0) or 0)
            resale = float(cells.get(3, 0) or 0)
            if recycle > 0 or resale > 0:
                sheet_data[model] = {"recycle": recycle, "resale": resale}
        data[sheet_name] = sheet_data
    wb.close()
    return data

def main():
    print("Extracting old prices...")
    old_data = extract_prices(OLD_FILE)
    print("Extracting new prices...")
    new_data = extract_prices(NEW_FILE)
    
    print("\n--- Price Comparison (4.21 vs 5.07) ---")
    for category in SHEETS_TO_CHECK:
        old_cat = old_data.get(category, {})
        new_cat = new_data.get(category, {})
        
        up_count = 0
        down_count = 0
        same_count = 0
        
        total_up_amount = 0
        total_down_amount = 0
        
        added_count = 0
        removed_count = 0
        
        for model, new_prices in new_cat.items():
            if model in old_cat:
                old_prices = old_cat[model]
                # We'll compare recycle price to gauge the trend
                diff = new_prices["recycle"] - old_prices["recycle"]
                
                if diff > 0:
                    up_count += 1
                    total_up_amount += diff
                elif diff < 0:
                    down_count += 1
                    total_down_amount += abs(diff)
                else:
                    same_count += 1
            else:
                added_count += 1
                
        for model in old_cat:
            if model not in new_cat:
                removed_count += 1
                    
        print(f"\n[{category}]:")
        print(f"  Total compared (exist in both): {up_count + down_count + same_count}")
        print(f"  Added: {added_count} items")
        print(f"  Removed: {removed_count} items")
        print(f"  Increased: {up_count} items (Avg increase: +{total_up_amount/up_count:.2f}元)" if up_count > 0 else "  Increased: 0")
        print(f"  Decreased: {down_count} items (Avg decrease: -{total_down_amount/down_count:.2f}元)" if down_count > 0 else "  Decreased: 0")
        print(f"  Unchanged: {same_count} items")
        
        if up_count > down_count:
            trend = "Overall Trend: INCREASE (涨价为主)"
        elif down_count > up_count:
            trend = "Overall Trend: DECREASE (降价为主)"
        else:
            trend = "Overall Trend: FLAT (平稳)"
        print(f"  => {trend}")

if __name__ == "__main__":
    main()
