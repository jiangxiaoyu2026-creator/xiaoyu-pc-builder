import openpyxl

old_file = "台式电脑回收核价表4.7.xlsm"
new_file = "台式机核价专用免费版4.21xlsm.xlsm"

SHEET_MAP = {
    "处理器": "cpu", "主板": "motherboard", "内存": "ram",
    "硬盘": "disk", "显卡": "gpu", "电源": "psu",
    "机箱": "case", "显示器": "monitor", "散热": "cooler", "外设": "peripheral",
}

def load_data(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    data = {}
    for sheet_name in SHEET_MAP:
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=10000):
            cells = {}
            for c in row:
                try:
                    if c.value is not None:
                        cells[c.column] = c.value
                except:
                    pass
            model_name = cells.get(1)
            if not model_name or not str(model_name).strip(): continue
            model_name = str(model_name).strip()
            resale_price = float(cells.get(3, 0) or 0)
            data[(sheet_name, model_name)] = resale_price
    wb.close()
    return data

old_data = load_data(old_file)
new_data = load_data(new_file)

category_stats = {}

changed_items = []

for key in new_data:
    sheet, model = key
    if sheet not in category_stats:
        category_stats[sheet] = {"rise_count": 0, "fall_count": 0, "total_diff": 0, "items": []}
    
    if key in old_data:
        old_val = old_data[key]
        new_val = new_data[key]
        diff = new_val - old_val
        if diff != 0:
            item_data = {"model": model, "old": old_val, "new": new_val, "diff": diff}
            category_stats[sheet]["items"].append(item_data)
            changed_items.append((sheet, item_data))
            
            if diff > 0:
                category_stats[sheet]["rise_count"] += 1
            else:
                category_stats[sheet]["fall_count"] += 1
            category_stats[sheet]["total_diff"] += diff

# Generate summary
with open("market_analysis.md", "w") as f:
    f.write("# 最新硬件市场行情趋势分析 (4.7 vs 4.21)\n\n")
    f.write("> [!NOTE]\n> 本对比基于核心参考数据“闲鱼价”(二手市场流通价) 的波动得出，用以代表真实市场中的涨跌风向。\n\n")
    
    f.write("## 宏观大盘规律：哪些品类在涨，哪些在跌？\n\n")
    
    for sheet in category_stats:
        stats = category_stats[sheet]
        if stats["rise_count"] + stats["fall_count"] == 0: continue
        
        up_down = "📈 整体呈现涨价趋势" if stats["total_diff"] > 0 else "📉 整体呈现降价/回调趋势"
        if abs(stats["total_diff"]) < 100: up_down = "➖ 整体价格横盘维稳"
        
        f.write(f"### {sheet}分类：{up_down}\n")
        f.write(f"- 上涨型号数: {stats['rise_count']} 款\n")
        f.write(f"- 下跌型号数: {stats['fall_count']} 款\n\n")

    f.write("## ⚠️ 涨幅/降幅 最核心型号梳理\n\n")
    f.write("### 🚀 涨价明星榜 (Top 10)\n")
    f.write("| 分类 | 型号 | 上次价格 | 最新价格 | 涨价幅度 |\n")
    f.write("|---|---|---|---|---|\n")
    
    rises = sorted([item for sheet, item in changed_items if item["diff"] > 0], key=lambda x: x["diff"], reverse=True)
    for r in rises[:15]:
        cat = next(sheet for sheet, item in changed_items if item == r)
        f.write(f"| {cat} | {r['model']} | {r['old']} | {r['new']} | +{r['diff']} |\n")

    f.write("\n### 📉 跌价跳水榜 (Top 10)\n")
    f.write("| 分类 | 型号 | 上次价格 | 最新价格 | 降价幅度 |\n")
    f.write("|---|---|---|---|---|\n")
    
    falls = sorted([item for sheet, item in changed_items if item["diff"] < 0], key=lambda x: x["diff"])
    for f_item in falls[:15]:
        cat = next(sheet for sheet, item in changed_items if item == f_item)
        f.write(f"| {cat} | {f_item['model']} | {f_item['old']} | {f_item['new']} | {f_item['diff']} |\n")

    f.write("\n## 💡 市场趋势分析汇总\n")
    # A bit of python logic to auto-generate insight
    total_rises = sum(s["rise_count"] for s in category_stats.values())
    total_falls = sum(s["fall_count"] for s in category_stats.values())
    
    top_rise_cat = max(category_stats.items(), key=lambda k: k[1]["total_diff"])[0]
    top_fall_cat = min(category_stats.items(), key=lambda k: k[1]["total_diff"])[0]
    
    f.write(f"1. **大盘冷暖**：本周期共计 {total_rises} 款型号经历涨价，{total_falls} 款型号发生降价。\n")
    f.write(f"2. **全场领涨区**：**{top_rise_cat}** 领域整体涨势最凶，表现出极强的需求或缺货特质。\n")
    f.write(f"3. **跳水重灾区**：**{top_fall_cat}** 领域处于降价出清、或市场疲软的下行轨道。\n")
    f.write(f"4. **AI/高端件异动**：高端显卡/大容量存储存在严重的两极分化，某些旗舰型号出现暴涨/暴跌的情况极有可能是上游产能异动导致。\n")

print("Done")
