import sys
import openpyxl
from sqlmodel import Session, select
from datetime import datetime

import os
# Allow importing from server_py
sys.path.append(os.path.join(os.path.dirname(__file__), 'server_py'))
from server_py.db import engine
from server_py.models import RecyclingPrice

file_path = "回收数价格表.xlsm"

SHEET_MAP = {
    "处理器": "cpu", "主板": "motherboard", "内存": "ram",
    "硬盘": "disk", "显卡": "gpu", "电源": "psu",
    "机箱": "case", "显示器": "monitor", "散热": "cooler", "外设": "peripheral",
}

def import_excel():
    print(f"Reading {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Failed to load workbook: {e}")
        return

    total_new = 0
    total_updated = 0
    
    with Session(engine) as session:
        for sheet_name, category in SHEET_MAP.items():
            if sheet_name not in wb.sheetnames:
                print(f"Sheet {sheet_name} not found, skipping.")
                continue
            
            print(f"Importing {sheet_name} ({category})...")
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, max_row=10000):
                cells = {}
                for c in row:
                    try:
                        if c.value is not None:
                            cells[c.column] = c.value
                    except:
                        pass
                
                model_name = cells.get(1)
                if not model_name or not str(model_name).strip():
                    continue
                model_name = str(model_name).strip()
                
                recycle_price = float(cells.get(2, 0) or 0)
                resale_price = float(cells.get(3, 0) or 0)
                if recycle_price == 0 and resale_price == 0:
                    continue
                
                validity_text = str(cells.get(4, ""))
                validity = "active" if validity_text in ("一周内", "一月内", "半月内", "三天内", "长期有效") else "expired"
                
                live_col = 14 if sheet_name == "处理器" else 12
                live_price = float(cells.get(live_col, 0) or 0) if cells.get(live_col) else None
                new_price = float(cells.get(7, 0) or 0) if cells.get(7) else None
                
                updated_at = cells.get(5)
                if isinstance(updated_at, datetime):
                    updated_at = updated_at.isoformat()
                else:
                    updated_at = str(updated_at) if updated_at else datetime.utcnow().isoformat()
                
                existing = session.exec(
                    select(RecyclingPrice)
                    .where(RecyclingPrice.category == category)
                    .where(RecyclingPrice.model == model_name)
                ).first()
                
                if existing:
                    existing.recyclePrice = recycle_price
                    existing.resalePrice = resale_price
                    existing.livePrice = live_price
                    existing.newPrice = new_price
                    existing.validity = validity
                    existing.updatedAt = updated_at
                    existing.updatedBy = "admin"
                    session.add(existing)
                    total_updated += 1
                else:
                    session.add(RecyclingPrice(
                        category=category, model=model_name,
                        recyclePrice=recycle_price, resalePrice=resale_price,
                        livePrice=live_price, newPrice=new_price,
                        validity=validity, updatedAt=updated_at,
                        updatedBy="admin",
                        note=str(cells.get(10, "")) if cells.get(10) else None,
                        imageUrl=str(cells.get(11, "")) if cells.get(11) else None,
                    ))
                    total_new += 1
            
            session.commit()
    
    wb.close()
    print(f"Import complete: {total_new} new, {total_updated} updated.")

if __name__ == "__main__":
    import_excel()
