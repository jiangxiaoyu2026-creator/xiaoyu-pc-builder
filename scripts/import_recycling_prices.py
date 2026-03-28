"""
回收价格表 Excel 导入脚本
从 回收数价格表.xlsm 导入 ~11,000 条 SKU 到 recycling_prices 表

Usage:
    cd /Users/mac/new
    python3 scripts/import_recycling_prices.py
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from sqlmodel import Session, select
from server_py.db import engine, init_db
from server_py.models import RecyclingPrice
from datetime import datetime

# Sheet 名称 -> 内部品类代码
SHEET_CATEGORY_MAP = {
    "处理器": "cpu",
    "主板": "motherboard",
    "内存": "ram",
    "硬盘": "disk",
    "显卡": "gpu",
    "电源": "psu",
    "机箱": "case",
    "显示器": "monitor",
    "散热": "cooler",
    "外设": "peripheral",
}

# 有效期文本 -> validity 状态
def parse_validity(text: str) -> str:
    if not text:
        return "expired"
    text = str(text).strip()
    if text in ("一周内", "一月内", "半月内", "三天内"):
        return "active"
    return "expired"

def parse_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def parse_date(val) -> str:
    if val is None:
        return datetime.utcnow().isoformat()
    if isinstance(val, datetime):
        return val.isoformat()
    return str(val)

def import_excel(filepath: str):
    print(f"📂 Opening: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    init_db()
    
    total_new = 0
    total_updated = 0
    total_skipped = 0
    
    with Session(engine) as session:
        for sheet_name, category in SHEET_CATEGORY_MAP.items():
            if sheet_name not in wb.sheetnames:
                print(f"  ⚠️  Sheet '{sheet_name}' not found, skipping")
                continue
                
            ws = wb[sheet_name]
            sheet_new = 0
            sheet_updated = 0
            sheet_skipped = 0
            
            for row in ws.iter_rows(min_row=2, max_row=10000):
                # Extract cell values safely
                cells = {}
                for c in row:
                    try:
                        if c.value is not None:
                            cells[c.column] = c.value
                    except:
                        pass
                
                # Column A = model name (required)
                model_name = cells.get(1)
                if not model_name or not str(model_name).strip():
                    continue
                model_name = str(model_name).strip()
                
                # Column B = recycle price, C = resale price (both required)
                recycle_price = parse_float(cells.get(2))
                resale_price = parse_float(cells.get(3))
                if recycle_price == 0 and resale_price == 0:
                    sheet_skipped += 1
                    continue
                
                # Other columns
                validity_text = str(cells.get(4, "")) 
                updated_at = parse_date(cells.get(5))
                updated_by = str(cells.get(6, "")) if cells.get(6) else None
                new_price = parse_float(cells.get(7)) if cells.get(7) else None
                note = str(cells.get(10, "")) if cells.get(10) else None
                image_url = str(cells.get(11, "")) if cells.get(11) else None
                
                # Live price: column 12 for most sheets, column 14 for 处理器
                live_price_col = 14 if sheet_name == "处理器" else 12
                live_price = parse_float(cells.get(live_price_col)) if cells.get(live_price_col) else None
                
                validity = parse_validity(validity_text)
                
                # Upsert: check if (category, model) exists
                existing = session.exec(
                    select(RecyclingPrice)
                    .where(RecyclingPrice.category == category)
                    .where(RecyclingPrice.model == model_name)
                ).first()
                
                if existing:
                    existing.recyclePrice = recycle_price
                    existing.resalePrice = resale_price
                    existing.livePrice = live_price
                    existing.newPrice = new_price
                    existing.validity = validity
                    existing.updatedAt = updated_at
                    existing.updatedBy = updated_by
                    existing.note = note
                    existing.imageUrl = image_url
                    session.add(existing)
                    sheet_updated += 1
                else:
                    new_item = RecyclingPrice(
                        category=category,
                        model=model_name,
                        recyclePrice=recycle_price,
                        resalePrice=resale_price,
                        livePrice=live_price,
                        newPrice=new_price,
                        validity=validity,
                        updatedAt=updated_at,
                        updatedBy=updated_by,
                        note=note,
                        imageUrl=image_url,
                    )
                    session.add(new_item)
                    sheet_new += 1
            
            session.commit()
            print(f"  ✅ {sheet_name} ({category}): +{sheet_new} new, ~{sheet_updated} updated, _{sheet_skipped} skipped")
            total_new += sheet_new
            total_updated += sheet_updated
            total_skipped += sheet_skipped
    
    wb.close()
    print(f"\n🎉 Import complete! New: {total_new}, Updated: {total_updated}, Skipped: {total_skipped}")
    print(f"   Total in DB: {total_new + total_updated}")


if __name__ == "__main__":
    excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "回收数价格表.xlsm")
    if not os.path.exists(excel_path):
        print(f"❌ File not found: {excel_path}")
        sys.exit(1)
    import_excel(excel_path)
