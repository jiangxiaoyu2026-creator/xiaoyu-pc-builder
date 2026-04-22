import openpyxl

old_file = "台式电脑回收核价表4.7.xlsm"
new_file = "台式机核价专用免费版4.21xlsm.xlsm"

SHEET_MAP = {
    "处理器": "cpu", "主板": "motherboard", "内存": "ram",
    "硬盘": "disk", "显卡": "gpu", "电源": "psu",
    "机箱": "case", "显示器": "monitor", "散热": "cooler", "外设": "peripheral",
}

def load_data(file_path):
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    data = {}
    for sheet_name in SHEET_MAP:
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=10000):
            cells = {}
            for c in row:
                try:
                    if c.value is not None:
                        cells[c.column] = c.value
                except:
                    pass
            model_name = cells.get(1)
            if not model_name or not str(model_name).strip(): continue
            model_name = str(model_name).strip()
            
            recycle_price = float(cells.get(2, 0) or 0)
            resale_price = float(cells.get(3, 0) or 0)
            live_col = 14 if sheet_name == "处理器" else 12
            live_price = float(cells.get(live_col, 0) or 0) if cells.get(live_col) else 0

            data[(sheet_name, model_name)] = {
                "recycle": recycle_price,
                "resale": resale_price,
                "live": live_price
            }
    wb.close()
    return data

def main():
    old_data = load_data(old_file)
    new_data = load_data(new_file)
    
    new_items = []
    changed_items = []
    
    for key, new_val in new_data.items():
        if key not in old_data:
            new_items.append((key, new_val))
        else:
            old_val = old_data[key]
            if abs(old_val["recycle"] - new_val["recycle"]) > 0.01 or \
               abs(old_val["resale"] - new_val["resale"]) > 0.01 or \
               abs(old_val["live"] - new_val["live"]) > 0.01:
                changed_items.append((key, old_val, new_val))
                
    with open("diff_report.md", "w") as f:
        f.write(f"# 价格表更新对比报告 (4.7 vs 4.21)\n\n")
        f.write(f"- 新增型号: {len(new_items)}个\n")
        f.write(f"- 价格变动: {len(changed_items)}个\n\n")
        
        f.write("## 📌 新增型号部分展示 (最多展示50条)\n\n")
        f.write("| 类别 | 型号 | 回收价 | 闲鱼价 | 直播价 |\n")
        f.write("|---|---|---|---|---|\n")
        for (sheet, model), val in new_items[:50]:
            f.write(f"| {sheet} | {model} | {val['recycle']} | {val['resale']} | {val['live']} |\n")
            
        f.write("\n## 📈 价格变动部分展示 (波动较大的50条)\n\n")
        # 简单按波动绝对值排下序
        changed_items.sort(key=lambda x: abs(x[2]["resale"] - x[1]["resale"]), reverse=True)
        
        f.write("| 类别 | 型号 | 旧回收价 -> 新回收价 | 旧闲鱼价 -> 新闲鱼价 |\n")
        f.write("|---|---|---|---|\n")
        for (sheet, model), old_v, new_v in changed_items[:50]:
            f.write(f"| {sheet} | {model} | {old_v['recycle']} -> {new_v['recycle']} | {old_v['resale']} -> {new_v['resale']} |\n")
            
    print("Done generating diff_report.md")

if __name__ == '__main__':
    main()
