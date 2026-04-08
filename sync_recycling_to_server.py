"""
Sync recycling prices from Excel (4.7) to production server via API.
- New products: create via POST
- Existing products with price changes: update via PUT
"""
import openpyxl
import requests
import sys
from datetime import datetime

EXCEL_PATH = "/Users/mac/Documents/dayrichang/台式电脑回收核价表4.7(1).xlsm"
API_BASE = "https://www.diyxx.com/api/recycling-prices"

SHEET_MAP = {
    "处理器": "cpu", "主板": "motherboard", "内存": "ram",
    "硬盘": "disk", "显卡": "gpu", "电源": "psu",
    "机箱": "case", "显示器": "monitor", "散热": "cooler", "外设": "peripheral",
}

def login():
    resp = requests.post("https://www.diyxx.com/api/auth/login", json={
        "username": "xiaoyu", "password": "jiangxiaoyu119"
    })
    return resp.json()["access_token"]

def fetch_all_server_items(token, category):
    """Fetch all items for a category from the server."""
    items = []
    page = 1
    while True:
        resp = requests.get(f"{API_BASE}/admin", params={
            "category": category, "page": page, "page_size": 200
        }, headers={"Authorization": f"Bearer {token}"})
        data = resp.json()
        items.extend(data.get("items", []))
        total = data.get("total", 0)
        if page * 200 >= total:
            break
        page += 1
    return items

def parse_excel():
    """Parse all items from the Excel file."""
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    all_items = []
    
    for sheet_name, category in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Sheet '{sheet_name}' not found, skipping.")
            continue
        
        ws = wb[sheet_name]
        count = 0
        for row in ws.iter_rows(min_row=2, max_row=10000):
            cells = {}
            for c in row:
                try:
                    if c.value is not None:
                        cells[c.column] = c.value
                except:
                    pass
            
            model_name = cells.get(1)
            if not model_name or not str(model_name).strip():
                continue
            model_name = str(model_name).strip()
            
            recycle_price = float(cells.get(2, 0) or 0)
            resale_price = float(cells.get(3, 0) or 0)
            if recycle_price == 0 and resale_price == 0:
                continue
            
            validity_text = str(cells.get(4, ""))
            validity = "active" if validity_text in ("一周内", "一月内", "半月内", "三天内", "长期有效") else "expired"
            
            # CPU uses col 14 for live price, others use col 12
            live_col = 14 if sheet_name == "处理器" else 12
            live_price = float(cells.get(live_col, 0) or 0) if cells.get(live_col) else None
            new_price = float(cells.get(7, 0) or 0) if cells.get(7) else None
            
            updated_at = cells.get(5)
            if isinstance(updated_at, datetime):
                updated_at = updated_at.isoformat()
            else:
                updated_at = str(updated_at) if updated_at else datetime.utcnow().isoformat()
            
            note = str(cells.get(10, "")) if cells.get(10) else None
            image_url = str(cells.get(11, "")) if cells.get(11) else None
            
            all_items.append({
                "category": category,
                "sheet": sheet_name,
                "model": model_name,
                "recyclePrice": recycle_price,
                "resalePrice": resale_price,
                "livePrice": live_price,
                "newPrice": new_price,
                "validity": validity,
                "updatedAt": updated_at,
                "note": note,
                "imageUrl": image_url,
            })
            count += 1
        print(f"  {sheet_name} ({category}): {count} items parsed")
    
    wb.close()
    return all_items

def main():
    print("=== Recycling Price Sync: Excel 4.7 -> Production ===\n")
    
    # 1. Parse Excel
    print("[1/4] Parsing Excel...")
    excel_items = parse_excel()
    print(f"\nTotal items in Excel: {len(excel_items)}\n")
    
    # 2. Login
    print("[2/4] Logging in to production API...")
    token = login()
    headers = {"Authorization": f"Bearer {token}"}
    print("  Login OK.\n")
    
    # 3. Fetch all server data and build lookup
    print("[3/4] Fetching current server data...")
    server_lookup = {}  # (category, model) -> server_item
    for sheet_name, category in SHEET_MAP.items():
        items = fetch_all_server_items(token, category)
        for item in items:
            key = (item.get("category", category), item.get("model", ""))
            server_lookup[key] = item
        print(f"  {sheet_name}: {len(items)} items on server")
    print(f"  Total on server: {len(server_lookup)}\n")
    
    # 4. Diff and sync
    print("[4/4] Syncing changes...")
    new_count = 0
    updated_count = 0
    skipped_count = 0
    errors = []
    
    for excel_item in excel_items:
        key = (excel_item["category"], excel_item["model"])
        server_item = server_lookup.get(key)
        
        if server_item is None:
            # NEW item - create it
            payload = {
                "category": excel_item["category"],
                "model": excel_item["model"],
                "recyclePrice": excel_item["recyclePrice"],
                "resalePrice": excel_item["resalePrice"],
                "livePrice": excel_item["livePrice"],
                "newPrice": excel_item["newPrice"],
                "validity": excel_item["validity"],
                "note": excel_item["note"],
                "imageUrl": excel_item["imageUrl"],
            }
            resp = requests.post(f"{API_BASE}/admin", json=payload, headers=headers)
            if resp.status_code in (200, 201):
                new_count += 1
                if new_count <= 10:
                    print(f"  + NEW: [{excel_item['sheet']}] {excel_item['model']} (回收:{excel_item['recyclePrice']} 闲鱼:{excel_item['resalePrice']})")
            else:
                errors.append(f"CREATE FAIL [{excel_item['model']}]: {resp.status_code} {resp.text[:100]}")
        else:
            # Check if prices changed
            s_recycle = float(server_item.get("recyclePrice", 0) or 0)
            s_resale = float(server_item.get("resalePrice", 0) or 0)
            s_live = float(server_item.get("livePrice", 0) or 0)
            e_live = float(excel_item.get("livePrice", 0) or 0)
            
            if (abs(s_recycle - excel_item["recyclePrice"]) > 0.01 or
                abs(s_resale - excel_item["resalePrice"]) > 0.01 or
                abs(s_live - e_live) > 0.01):
                # Price changed - update
                item_id = server_item.get("id")
                payload = {
                    "recyclePrice": excel_item["recyclePrice"],
                    "resalePrice": excel_item["resalePrice"],
                    "livePrice": excel_item["livePrice"],
                    "newPrice": excel_item["newPrice"],
                    "validity": excel_item["validity"],
                }
                resp = requests.put(f"{API_BASE}/admin/{item_id}", json=payload, headers=headers)
                if resp.status_code == 200:
                    updated_count += 1
                    if updated_count <= 10:
                        print(f"  ~ UPDATE: [{excel_item['sheet']}] {excel_item['model']} "
                              f"(回收: {s_recycle}->{excel_item['recyclePrice']}, "
                              f"闲鱼: {s_resale}->{excel_item['resalePrice']})")
                else:
                    errors.append(f"UPDATE FAIL [{excel_item['model']}]: {resp.status_code} {resp.text[:100]}")
            else:
                skipped_count += 1
    
    # Summary
    print(f"\n{'='*50}")
    print(f"Sync Complete!")
    print(f"  New items added:   {new_count}")
    print(f"  Items updated:     {updated_count}")
    print(f"  Items unchanged:   {skipped_count}")
    if new_count > 10:
        print(f"  (only first 10 new/updated items shown above)")
    if errors:
        print(f"\n  ERRORS ({len(errors)}):")
        for e in errors[:20]:
            print(f"    {e}")

if __name__ == "__main__":
    main()
